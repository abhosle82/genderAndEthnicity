# -*- coding: utf-8 -*-
"""
Created on Sat May 14 12:23:13 2022

@author: uma
"""
import pandas as pd
# from ethnicolr import census_ln, pred_census_ln
import gender_guesser.detector as gender
import pandas as pd
import math

gd_detector = gender.Detector(case_sensitive=False)

# print(gd_detector.get_gender(u"Uma"))


def get_first_name(name):
    name = name.split('-')
    name = name[0].split(' ')
    return name[0]


def get_gender(name):
    gender = gd_detector.get_gender(name)
    if "male" == gender:
        return 0
    else:
        return 1


# excel_file = 'C:\Hackathon\HackathonDataMinorityWomenOwned2022v1.xlsx'
# excel_file = 'C:\Users\umack\wfhackathon\HackathonDataMinorityWomenOwned2022v1.xlsx'
df = pd.read_excel(r"C:\Users\umack\wfhackathon\hackathon_flname_Leader_2.xlsx")
print(get_gender(df['FirstName'].iloc[1]))
isWomanOwned = []

# get the iswomanowned data using gender-guesser model
for index, row in df.iterrows():
    # print(index,row)
    contact = str(row["FirstName"])
    # print(contact)
    if pd.isnull(contact):
        contact = row["FirstName"]

        isWomanOwned.append(get_gender((contact)))
    else:

        isWomanOwned.append(get_gender((contact)))

from openpyxl import load_workbook

df_new = pd.DataFrame({'gender': isWomanOwned})
wb = load_workbook(r"C:\Users\umack\wfhackathon\hackathon_flname_Leader_2.xlsx")

ws = wb['DiversityInclusionData1']

for index, row in df_new.iterrows():
    cell = 'D%d' % (index + 2)
    ws[cell] = row[0]

wb.save(r"C:\Users\umack\wfhackathon\hackathon_flname_Leader_2.xlsx")

'''for index, row in df.iterrows():
    contact = str(row["executiveContact1"])
    print(contact)
    if pd.isnull(contact):
        contact = row["executiveContact2"]
        print(get_gender(get_first_name(contact)))
    else:
        print(get_gender(get_first_name(contact)))'''
